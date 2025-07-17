import os
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import traceback

def process_excel_files():
    folder_path = filedialog.askdirectory(title="Excelファイルが格納されているフォルダを選択してください")
    if not folder_path:
        return

    excel_files = [f for f in os.listdir(folder_path) if f.endswith('.xlsx')]
    
    if not excel_files:
        messagebox.showwarning("警告", "選択されたフォルダにExcelファイルが見つかりません。")
        return

    for file_name in excel_files:
        file_path = os.path.join(folder_path, file_name)
        try:
            process_single_file(file_path)
        except Exception as e:
            error_msg = f"ファイル {file_name} の処理中にエラーが発生しました:\n{str(e)}\n\n"
            error_msg += "詳細なエラー情報:\n" + traceback.format_exc()
            messagebox.showerror("エラー", error_msg)

    messagebox.showinfo("成功", f"{len(excel_files)}個のExcelファイルが処理されました。")

def process_single_file(file_path):
    wb = load_workbook(file_path)
    source_sheet = wb.worksheets[-1]
    new_sheet_name = sheet_name_entry.get()
    
    if new_sheet_name in wb.sheetnames:
        raise ValueError(f"シート名 '{new_sheet_name}' は既に存在します。")
    
    new_sheet = wb.copy_worksheet(source_sheet)
    new_sheet.title = new_sheet_name

    # セルの結合を確実にコピー
    new_sheet.merged_cells = source_sheet.merged_cells

    # 列の幅をコピー
    for column in range(1, source_sheet.max_column + 1):
        column_letter = get_column_letter(column)
        new_sheet.column_dimensions[column_letter].width = source_sheet.column_dimensions[column_letter].width

    # 行の高さをコピー
    for row in range(1, source_sheet.max_row + 1):
        new_sheet.row_dimensions[row].height = source_sheet.row_dimensions[row].height

    # データ検証（選択タブ）をコピー
    for validation in source_sheet.data_validations.dataValidation:
        new_sheet.add_data_validation(validation)

    # 条件付き書式をコピー
    for rule in source_sheet.conditional_formatting:
        new_sheet.conditional_formatting.add(rule.cells.coord, rule.cf)

    try:
        # B2:C2 に日付を入力
        date_value = date_entry.get()
        new_sheet['B2'] = date_value
        new_sheet['C2'] = date_value

        # C3とC4に出勤時間を入力
        start_time = start_time_entry.get()
        new_sheet['C3'] = start_time
        new_sheet['C4'] = start_time

        # E3とE4に退勤時間を入力
        end_time = end_time_entry.get()
        new_sheet['E3'] = end_time
        new_sheet['E4'] = end_time

        # H2に体温を入力（選択タブを維持）
        temperature = temperature_var.get()
        new_sheet['H2'] = temperature

        # B5:E5に睡眠状態を入力（選択タブを維持）
        sleep_quality = sleep_quality_var.get()
        for cell in ['B5', 'C5', 'D5', 'E5']:
            new_sheet[cell] = sleep_quality

        # H5:J5に気分状態を入力（選択タブを維持）
        mood = mood_var.get()
        for cell in ['H5', 'I5', 'J5']:
            new_sheet[cell] = mood

        # G3に就寝時刻を入力
        bedtime = bedtime_entry.get()
        new_sheet['G3'] = bedtime

        # I3:J3に起床時刻を入力
        wake_time = wake_time_entry.get()
        new_sheet['I3'] = wake_time
        new_sheet['J3'] = wake_time

        # B8:J11に振り返り・感想を入力
        reflection = reflection_text.get("1.0", tk.END).strip()
        new_sheet['B8'] = reflection

        # J2に症状を入力
        symptom = symptom_entry.get()
        new_sheet['J2'] = symptom

        # H4に服薬状況を入力
        medication = medication_entry.get()
        new_sheet['H4'] = medication

        # B12:J16 の範囲のセルをクリア
        for row in new_sheet['B12:J16']:
            for cell in row:
                cell.value = None

    except Exception as e:
        raise Exception(f"セルの値設定中にエラーが発生しました: {str(e)}")

    try:
        wb.save(file_path)
    except PermissionError:
        raise Exception(f"ファイル '{file_path}' への保存権限がありません。ファイルが開いていないか確認してください。")
    except Exception as e:
        raise Exception(f"ファイルの保存中にエラーが発生しました: {str(e)}")

# GUIの作成
root = tk.Tk()
root.title("個人日報記入ツール")

tk.Label(root, text="新しいシート名:").grid(row=0, column=0, padx=5, pady=5, sticky="e")
sheet_name_entry = tk.Entry(root)
sheet_name_entry.grid(row=0, column=1, padx=5, pady=5, sticky="w")
sheet_name_entry.insert(0, datetime.now().strftime("%Y%m%d"))

tk.Label(root, text="日付 (YYYY/MM/DD):").grid(row=1, column=0, padx=5, pady=5, sticky="e")
date_entry = tk.Entry(root)
date_entry.grid(row=1, column=1, padx=5, pady=5, sticky="w")
date_entry.insert(0, datetime.now().strftime("%Y/%m/%d"))

tk.Label(root, text="出勤時間:").grid(row=2, column=0, padx=5, pady=5, sticky="e")
start_time_entry = tk.Entry(root)
start_time_entry.grid(row=2, column=1, padx=5, pady=5, sticky="w")
start_time_entry.insert(0, "09:00")

tk.Label(root, text="退勤時間:").grid(row=3, column=0, padx=5, pady=5, sticky="e")
end_time_entry = tk.Entry(root)
end_time_entry.grid(row=3, column=1, padx=5, pady=5, sticky="w")
end_time_entry.insert(0, "12:00")

tk.Label(root, text="体温:").grid(row=4, column=0, padx=5, pady=5, sticky="e")
temperature_var = tk.StringVar(root)
temperature_choices = [f"{i/10:.1f}" for i in range(357, 373)]  # 35.7から37.2まで
temperature_menu = ttk.Combobox(root, textvariable=temperature_var, values=temperature_choices)
temperature_menu.grid(row=4, column=1, padx=5, pady=5, sticky="w")
temperature_menu.set("36.5")  # デフォルト値

tk.Label(root, text="睡眠状態:").grid(row=5, column=0, padx=5, pady=5, sticky="e")
sleep_quality_var = tk.StringVar(root)
sleep_quality_choices = ["良く眠れた", "眠れた", "あまり眠れなかった", "眠れなかった"]
sleep_quality_menu = ttk.Combobox(root, textvariable=sleep_quality_var, values=sleep_quality_choices)
sleep_quality_menu.grid(row=5, column=1, padx=5, pady=5, sticky="w")
sleep_quality_menu.set("眠れた")  # デフォルト値

tk.Label(root, text="気分状態:").grid(row=6, column=0, padx=5, pady=5, sticky="e")
mood_var = tk.StringVar(root)
mood_choices = ["落ち着いている", "不安感がある", "少し落ち込んでいる", "落ち着かない", "気分が重い", "スッキリしない"]
mood_menu = ttk.Combobox(root, textvariable=mood_var, values=mood_choices)
mood_menu.grid(row=6, column=1, padx=5, pady=5, sticky="w")
mood_menu.set("落ち着いている")  # デフォルト値

tk.Label(root, text="就寝時刻:").grid(row=7, column=0, padx=5, pady=5, sticky="e")
bedtime_entry = tk.Entry(root)
bedtime_entry.grid(row=7, column=1, padx=5, pady=5, sticky="w")

tk.Label(root, text="起床時刻:").grid(row=8, column=0, padx=5, pady=5, sticky="e")
wake_time_entry = tk.Entry(root)
wake_time_entry.grid(row=8, column=1, padx=5, pady=5, sticky="w")

tk.Label(root, text="振り返り・感想:").grid(row=9, column=0, padx=5, pady=5, sticky="ne")
reflection_text = scrolledtext.ScrolledText(root, width=40, height=5)
reflection_text.grid(row=9, column=1, padx=5, pady=5, sticky="w")

tk.Label(root, text="症状:").grid(row=10, column=0, padx=5, pady=5, sticky="e")
symptom_entry = tk.Entry(root)
symptom_entry.grid(row=10, column=1, padx=5, pady=5, sticky="w")
symptom_entry.insert(0, "無")  # デフォルト値

tk.Label(root, text="服薬状況変更:").grid(row=11, column=0, padx=5, pady=5, sticky="e")
medication_entry = tk.Entry(root)
medication_entry.grid(row=11, column=1, padx=5, pady=5, sticky="w")
medication_entry.insert(0, "無し")  # デフォルト値

process_button = tk.Button(root, text="フォルダを選択してExcelファイルを処理", command=process_excel_files)
process_button.grid(row=12, column=0, columnspan=2, pady=20)

root.mainloop()